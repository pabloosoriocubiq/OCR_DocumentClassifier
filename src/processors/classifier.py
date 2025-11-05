from typing import Tuple, List, Dict
from pathlib import Path
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import re
import os
from src.utils.logger import Logger
from src.config import DOCUMENT_TYPES, CLASSIFICATION_FOLDER


class DocumentClassifier:
    
    def __init__(self):
        self.logger = Logger.get_logger(__name__)
        self.document_types = DOCUMENT_TYPES
    
    def find_keywords_smart(self, text: str, keywords: List[str]) -> List[str]:

        text_lower = text.lower()
        found = []

        sorted_keywords = sorted(keywords, key=len, reverse=True)
        
        for keyword in sorted_keywords:
            keyword_lower = keyword.lower()
            
            #patrón word boundaries
            pattern = r'\b' + re.escape(keyword_lower) + r'\b'
            
            if re.search(pattern, text_lower):
                found.append(keyword)
        
        return found
    
    def classify_page(self, text: str) -> Tuple[str, List[str], List[str], int, int]:
        
        candidates = {}
        
        for doc_type, config in self.document_types.items():
            
            primary_keywords = config.get('primary_keywords', config.get('keywords', []))
            primary_found = self.find_keywords_smart(text, primary_keywords)
            
            if not primary_found:
                continue
            
            secondary_keywords = config.get('secondary_keywords', [])
            secondary_found = self.find_keywords_smart(text, secondary_keywords)

            is_functional = config.get('functional', False)
            min_secondary = config.get('min_secondary_matches', 0)
            
            if is_functional and len(secondary_found) < min_secondary:
                continue

            score = (len(primary_found) * 3) + len(secondary_found)
            if not is_functional:
                score = score *0.7
              
            total_keywords = len(primary_found) + len(secondary_found)
            
            candidates[doc_type] = {
                'score': score,
                'primary': primary_found,
                'secondary': secondary_found,
                'total_keywords': total_keywords,
                'functional': is_functional
            }
            
        
        if not candidates:
            return "UNKNOWN", [], [], 0, 0 

        def selection_criteria(doc_type):
            return (
            candidates[doc_type]['score'],
            candidates[doc_type]['functional'],
            candidates[doc_type]['total_keywords']
    )

        winner_type = max(candidates, key=selection_criteria)
        winner_data = candidates[winner_type]     
        num_candidates = len(candidates)
        
        return (
            winner_type,
            winner_data['primary'],
            winner_data['secondary'],
            winner_data['total_keywords'],
            num_candidates
        )
    
    def is_functional(self, doc_type: str) -> bool:
        if doc_type == "UNKNOWN":
            return False
        return self.document_types.get(doc_type, {}).get('functional', False)
    
    def group_consecutive_pages(self, classifications: List[Dict]) -> List[Dict]:
        groups = []
        current_group = None
        
        for page_class in classifications:
            if not page_class['functional']:
                current_group = None
                continue
            
            doc_type = page_class['document_type']
            page_num = page_class['page_number']
            
            if current_group is None or current_group['type'] != doc_type:
                current_group = {
                    'type': doc_type,
                    'pages': [page_num],
                    'start_page': page_num,
                    'end_page': page_num
                }
                groups.append(current_group)
            else:
                current_group['pages'].append(page_num)
                current_group['end_page'] = page_num
        
        return groups
    
    def save_classification_report(self, 
                                   pdf_name: str,
                                   classifications: List[Dict],
                                   groups: List[Dict]) -> Path:

        pdf_stem = Path(pdf_name).stem
        output_folder = CLASSIFICATION_FOLDER / pdf_stem
        output_folder.mkdir(parents=True, exist_ok=True)
        
        functional_count = sum(1 for c in classifications if c['functional'])
        non_functional_count = len(classifications) - functional_count
        
        report_data = {
            'pdf_name': pdf_name,
            'total_pages': len(classifications),
            'functional_pages': functional_count,
            'non_functional_pages': non_functional_count,
            'classifications': classifications,
            'document_groups': groups
        }
        
        # Guardar JSON
        json_path = output_folder / f"{pdf_stem}_clasificacion.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)
        
        # Guardar TXT
        txt_path = output_folder / f"{pdf_stem}_reporte.txt"
        self._generate_text_report(report_data, txt_path)
        
        return json_path
    
    def _generate_text_report(self, report_data: Dict, report_path: Path):
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("="*70 + "\n")
            f.write("REPORTE DE CLASIFICACIÓN\n")
            f.write("="*70 + "\n\n")
            
            f.write(f"Documento: {report_data['pdf_name']}\n")
            f.write(f"Total páginas: {report_data['total_pages']}\n")
            f.write(f"Funcionales: {report_data['functional_pages']}\n")
            f.write(f"Eliminadas: {report_data['non_functional_pages']}\n\n")
            
            f.write("-"*70 + "\n")
            f.write("CLASIFICACIÓN POR PÁGINA\n")
            f.write("-"*70 + "\n\n")
            
            for page in report_data['classifications']:
                status = "✓ MANTENER" if page['functional'] else "✗ ELIMINAR"
                f.write(f"Página {page['page_number']}: {page['document_type']} - {status}\n")
                keywords_str = ', '.join(page['keywords_found'])
                f.write(f"  Keywords: {keywords_str}\n\n")
            
            f.write("-"*70 + "\n")
            f.write("DOCUMENTOS DETECTADOS\n")
            f.write("-"*70 + "\n\n")
            
            for i, group in enumerate(report_data['document_groups'], 1):
                f.write(f"{i}. {group['type']}\n")
                f.write(f"   Páginas: {group['start_page']} - {group['end_page']}\n")
                f.write(f"   Total: {len(group['pages'])} páginas\n\n")
                
    def save_classification_excel(self, 
                              pdf_path: str,
                              pdf_name: str,
                              classifications: List[Dict],
                              excel_path: str = "reporte_clasificacion.xlsx"):
        MAX_PAGES = 10

        row_data = {'Path': pdf_path, 'PDF_Name': pdf_name}
        for i in range(1, MAX_PAGES):
            if i <= len(classifications):
                page = classifications[i-1]
                doc_type = page.get('document_type', 'UNKNOWN')
                is_func = page.get('functional', False)
                conf = page.get('ocr_confidence', 0.0)
                
                row_data[f'Pag_{i}'] = f"{doc_type} (x)" if (not is_func and doc_type != "UNKNOWN") else doc_type
                row_data[f'Confidence_{i}'] = f"{conf * 100:.1f}%" 
            else:
                row_data[f'Pag_{i}'] = ""
                row_data[f'Confidence_{i}'] = ""
        
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            work_sheet = wb.active
        else:
            wb = Workbook()
            work_sheet = wb.active
            work_sheet.title = "reporte_clasificacion"
            
            headers = ['Path', 'PDF_Name']
            for i in range(1, MAX_PAGES + 1):
                headers.extend([f'Pag {i}', f'Confidence {i}'])
            work_sheet.append(headers)
            
            for cell in work_sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                           
        row_values = [row_data['Path'], row_data['PDF_Name']]
        for i in range(1, MAX_PAGES + 1):
            row_values.extend([row_data[f'Pag_{i}'], row_data[f'Confidence_{i}']])
        work_sheet.append(row_values)

        work_sheet.column_dimensions['A'].width = 40
        work_sheet.column_dimensions['B'].width = 25
        for i in range(MAX_PAGES):
            work_sheet.column_dimensions[chr(67 + i*2)].width = 21 
            work_sheet.column_dimensions[chr(68 + i*2)].width = 13
            
        for row in work_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center')   
        
        wb.save(excel_path)
        self.logger.info(f" ✓ Excel actualizado: {excel_path}")